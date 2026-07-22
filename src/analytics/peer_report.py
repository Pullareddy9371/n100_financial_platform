import sqlite3
import pandas as pd
import os

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

DB_PATH = "db/nifty100.db"


class PeerReport:

    def __init__(self):
        self.conn = sqlite3.connect(DB_PATH)

    def load_data(self):

        peer = pd.read_sql(
            "SELECT * FROM peer_percentiles",
            self.conn
        )

        companies = pd.read_sql(
            """
            SELECT
                id,
                company_name
            FROM companies
            """,
            self.conn
        )

        df = peer.merge(
            companies,
            left_on="company_id",
            right_on="id",
            how="left"
        )

        return df

    def export_excel(self, df):

        os.makedirs("src/output", exist_ok=True)

        output_file = "src/output/peer_comparison.xlsx"

        # Write Excel
        with pd.ExcelWriter(output_file, engine="openpyxl") as writer:

            sectors = sorted(df["broad_sector"].dropna().unique())

            for sector in sectors:

                temp = df[
                    df["broad_sector"] == sector
                ].copy()

                temp.to_excel(
                    writer,
                    sheet_name=sector[:31],
                    index=False
                )

        # Open workbook for formatting
        wb = load_workbook(output_file)

        green = PatternFill(
            fill_type="solid",
            fgColor="90EE90"
        )

        yellow = PatternFill(
            fill_type="solid",
            fgColor="FFF59D"
        )

        red = PatternFill(
            fill_type="solid",
            fgColor="FFB6C1"
        )

        for ws in wb.worksheets:

            headers = []

            for cell in ws[1]:
                headers.append(cell.value)

            if "percentile_rank" not in headers:
                continue

            col = headers.index("percentile_rank") + 1

            for r in range(2, ws.max_row + 1):

                cell = ws.cell(row=r, column=col)

                try:
                    value = float(cell.value)
                except:
                    continue

                if value >= 75:
                    cell.fill = green

                elif value >= 25:
                    cell.fill = yellow

                else:
                    cell.fill = red

        wb.save(output_file)

        print("\npeer_comparison.xlsx created successfully!")

    def run(self):

        df = self.load_data()

        self.export_excel(df)

        print(df.head())


if __name__ == "__main__":

    PeerReport().run()